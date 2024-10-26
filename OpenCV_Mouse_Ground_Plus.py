import matplotlib.pyplot as plt
from scipy.ndimage import gaussian_filter
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import numpy as np
import cv2
import math
import time

start = time.time()

def excel_y(start_level, start_grid_y, grid_positon_y, dy):
    start_level = start_level
    grid_positon_y = grid_positon_y
    dy = dy
    start_grid_y = start_grid_y
    grid_to_elevation = []
    num = np.size(grid_find_position_y)
    for i in range(0, num):
        if i % 2 == 0:
            grid_to_elevation.append(start_level - ((grid_positon_y[i] - start_grid_y) / (dy + dy2)) * 2)
        elif i % 2 == 1:
            grid_to_elevation.append(start_level - ((grid_positon_y[i] - start_grid_y - dy2) / (dy + dy2)) * 2 - 1)
    return grid_to_elevation
def excel_x(start_level, start_grid_x, grid_positon_x, dx):
    start_level = start_level
    grid_positon_x = grid_positon_x
    dx = dx
    start_grid_x = start_grid_x
    grid_to_elevation = []
    num = np.size(grid_find_position_x)
    for i in range(0, num):
        if i % 2 == 0:
            grid_to_elevation.append(start_level + ((grid_positon_x[i] - start_grid_x) / (dx + dx2)) * 4)
        elif i % 2 == 1:
            grid_to_elevation.append(start_level + ((grid_positon_x[i] - start_grid_x - dx2) / (dx + dx2)) * 4 + 2)
    return grid_to_elevation
def Left_road_design(center_x, center_y, type, kerb, curb, drain_cover):
    Left_road_design_x = []
    Left_road_design_y = []
    if type == 'TYPEA':
        if divisional_island[road_i] == 0:
            center_y = center_y - kerb
            Left_road_design_x.append(center_x)
            Left_road_design_y.append(center_y)
            Left_road_design_x.append(center_x - divisional_island[road_i])
            Left_road_design_y.append(center_y)
            Left_road_design_x.append(Left_road_design_x[-1] - car_lane[road_i])
            Left_road_design_y.append(Left_road_design_y[-1] - car_lane[road_i] * 0.02)
            Left_road_design_x.append(Left_road_design_x[-1] - drain[road_i])
            Left_road_design_y.append(Left_road_design_y[-1])
            Left_road_design_x.append(Left_road_design_x[-1])
            Left_road_design_y.append(Left_road_design_y[-1] + drain_cover)
            Left_road_design_x.append(Left_road_design_x[-1] - sidewalk[road_i])
            Left_road_design_y.append(Left_road_design_y[-1] + sidewalk[road_i] * 0.02)
            Left_road_design = np.vstack((Left_road_design_y, Left_road_design_x))
        else:
            Left_road_design_x.append(center_x)
            Left_road_design_y.append(center_y)
            Left_road_design_x.append(center_x - divisional_island[road_i])
            Left_road_design_y.append(center_y)
            Left_road_design_x.append(Left_road_design_x[-1])
            Left_road_design_y.append(center_y - kerb)
            Left_road_design_x.append(Left_road_design_x[-1] - car_lane[road_i])
            Left_road_design_y.append(Left_road_design_y[-1] - car_lane[road_i] * 0.02)
            Left_road_design_x.append(Left_road_design_x[-1] - drain[road_i])
            Left_road_design_y.append(Left_road_design_y[-1])
            Left_road_design_x.append(Left_road_design_x[-1])
            Left_road_design_y.append(Left_road_design_y[-1] + drain_cover)
            Left_road_design_x.append(Left_road_design_x[-1] - sidewalk[road_i])
            Left_road_design_y.append(Left_road_design_y[-1] + sidewalk[road_i] * 0.02)
            Left_road_design = np.vstack((Left_road_design_y, Left_road_design_x))
    elif type == 'TYPEB':
        Left_road_design_x.append(center_x)
        Left_road_design_y.append(center_y)
        Left_road_design_x.append(center_x - divisional_island[road_i])
        Left_road_design_y.append(center_y)
        Left_road_design_x.append(Left_road_design_x[-1])
        Left_road_design_y.append(center_y - kerb)
        Left_road_design_x.append(Left_road_design_x[-1] - car_lane[road_i])
        Left_road_design_y.append(Left_road_design_y[-1] - car_lane[road_i] * 0.02)
        Left_road_design_x.append(Left_road_design_x[-1] - drain[road_i])
        Left_road_design_y.append(Left_road_design_y[-1])
        Left_road_design_x.append(Left_road_design_x[-1] - on_street_parking[road_i])
        Left_road_design_y.append(Left_road_design_y[-1] + on_street_parking[road_i] * 0.02)
        Left_road_design_x.append(Left_road_design_x[-1])
        Left_road_design_y.append(Left_road_design_y[-1] + curb)
        Left_road_design_x.append(Left_road_design_x[-1] - sidewalk[road_i])
        Left_road_design_y.append(Left_road_design_y[-1] + sidewalk[road_i] * 0.02)
        Left_road_design = np.vstack((Left_road_design_y, Left_road_design_x))
    elif type == 'TYPEC':
        center_y = center_y - kerb
        Left_road_design_x.append(center_x)
        Left_road_design_y.append(center_y)
        Left_road_design_x.append(center_x - divisional_island[road_i])
        Left_road_design_y.append(center_y)
        Left_road_design_x.append(Left_road_design_x[-1] - car_lane[road_i])
        Left_road_design_y.append(Left_road_design_y[-1] - car_lane[road_i] * 0.0137)
        Left_road_design_x.append(Left_road_design_x[-1] - drain[road_i])
        Left_road_design_y.append(Left_road_design_y[-1])
        Left_road_design_x.append(Left_road_design_x[-1])
        Left_road_design_y.append(Left_road_design_y[-1] + drain_cover)
        Left_road_design_x.append(Left_road_design_x[-1] - sidewalk[road_i] - guardrail)
        Left_road_design_y.append(Left_road_design_y[-1] + (sidewalk[road_i] + guardrail) * 0.02)
        Left_road_design = np.vstack((Left_road_design_y, Left_road_design_x))
    elif type == 'TYPED':
        center_y = center_y - kerb
        Left_road_design_x.append(center_x - divisional_island[road_i])
        Left_road_design_y.append(center_y)
        Left_road_design_x.append(Left_road_design_x[-1] - car_lane[road_i])
        Left_road_design_y.append(Left_road_design_y[-1] - car_lane[road_i] * 0.027)
        Left_road_design_x.append(Left_road_design_x[-1] - drain[road_i])
        Left_road_design_y.append(Left_road_design_y[-1])
        Left_road_design_x.append(Left_road_design_x[-1])
        Left_road_design_y.append(Left_road_design_y[-1] + drain_cover)
        Left_road_design_x.append(Left_road_design_x[-1] - sidewalk[road_i])
        Left_road_design_y.append(Left_road_design_y[-1] + sidewalk[road_i] * 0.027)
        Left_road_design = np.vstack((Left_road_design_y, Left_road_design_x))
    elif type == 'TYPEE':
        center_y = center_y - kerb
        Left_road_design_x.append(center_x)
        Left_road_design_y.append(center_y)
        Left_road_design_x.append(Left_road_design_x[-1] - car_lane[road_i])
        Left_road_design_y.append(Left_road_design_y[-1] - car_lane[road_i] * 0.02)
        Left_road_design_x.append(Left_road_design_x[-1] - sidewalk[road_i])
        Left_road_design_y.append(Left_road_design_y[-1])
        Left_road_design_x.append(Left_road_design_x[-1])
        Left_road_design_y.append(Left_road_design_y[-1] + drain_cover)
        Left_road_design = np.vstack((Left_road_design_y, Left_road_design_x))
    n = (np.shape(Left_road_design))[1]
    j = []
    for i in range(1, n):
        if Left_road_design[0][i - 1] == Left_road_design[0][i] and Left_road_design[1][i - 1] == Left_road_design[1][i]:
            j.append(i)
    Left_road_design = np.delete(Left_road_design, [j], 1)
    return Left_road_design
def Right_road_design(center_x, center_y, type, kerb, curb, drain_cover):
    Right_road_design_x = []
    Right_road_design_y = []
    if type == 'TYPEA':
        if divisional_island[road_i] == 0:
            center_y = center_y - kerb
            Right_road_design_x.append(center_x)
            Right_road_design_y.append(center_y)
            Right_road_design_x.append(center_x + divisional_island[road_i])
            Right_road_design_y.append(center_y)
            Right_road_design_x.append(Right_road_design_x[-1] + car_lane[road_i])
            Right_road_design_y.append(Right_road_design_y[-1] - car_lane[road_i] * 0.02)
            Right_road_design_x.append(Right_road_design_x[-1] + drain[road_i])
            Right_road_design_y.append(Right_road_design_y[-1])
            Right_road_design_x.append(Right_road_design_x[-1])
            Right_road_design_y.append(Right_road_design_y[-1] + drain_cover)
            Right_road_design_x.append(Right_road_design_x[-1] + sidewalk[road_i])
            Right_road_design_y.append(Right_road_design_y[-1] + sidewalk[road_i] * 0.02)
            Right_road_design = np.vstack((Right_road_design_y, Right_road_design_x))
        else:
            Right_road_design_x.append(center_x)
            Right_road_design_y.append(center_y)
            Right_road_design_x.append(center_x + divisional_island[road_i])
            Right_road_design_y.append(center_y)
            Right_road_design_x.append(Right_road_design_x[-1])
            Right_road_design_y.append(center_y - kerb)
            Right_road_design_x.append(Right_road_design_x[-1] + car_lane[road_i])
            Right_road_design_y.append(Right_road_design_y[-1] - car_lane[road_i] * 0.02)
            Right_road_design_x.append(Right_road_design_x[-1] + drain[road_i])
            Right_road_design_y.append(Right_road_design_y[-1])
            Right_road_design_x.append(Right_road_design_x[-1])
            Right_road_design_y.append(Right_road_design_y[-1] + drain_cover)
            Right_road_design_x.append(Right_road_design_x[-1] + sidewalk[road_i])
            Right_road_design_y.append(Right_road_design_y[-1] + sidewalk[road_i] * 0.02)
            Right_road_design = np.vstack((Right_road_design_y, Right_road_design_x))
    elif type == 'TYPEB':
        Right_road_design_x.append(center_x)
        Right_road_design_y.append(center_y)
        Right_road_design_x.append(center_x + divisional_island[road_i])
        Right_road_design_y.append(center_y)
        Right_road_design_x.append(Right_road_design_x[-1])
        Right_road_design_y.append(center_y - kerb)
        Right_road_design_x.append(Right_road_design_x[-1] + car_lane[road_i])
        Right_road_design_y.append(Right_road_design_y[-1] - car_lane[road_i] * 0.02)
        Right_road_design_x.append(Right_road_design_x[-1] + drain[road_i])
        Right_road_design_y.append(Right_road_design_y[-1])
        Right_road_design_x.append(Right_road_design_x[-1] + on_street_parking[road_i])
        Right_road_design_y.append(Right_road_design_y[-1] + on_street_parking[road_i] * 0.02)
        Right_road_design_x.append(Right_road_design_x[-1])
        Right_road_design_y.append(Right_road_design_y[-1] + curb)
        Right_road_design_x.append(Right_road_design_x[-1] + sidewalk[road_i])
        Right_road_design_y.append(Right_road_design_y[-1] + sidewalk[road_i] * 0.02)
        Right_road_design = np.vstack((Right_road_design_y, Right_road_design_x))
    elif type == 'TYPEC':
        center_y = center_y - kerb
        Right_road_design_x.append(center_x)
        Right_road_design_y.append(center_y)
        Right_road_design_x.append(center_x + divisional_island[road_i])
        Right_road_design_y.append(center_y)
        Right_road_design_x.append(Right_road_design_x[-1] + car_lane[road_i])
        Right_road_design_y.append(Right_road_design_y[-1] - car_lane[road_i] * 0.02)
        Right_road_design_x.append(Right_road_design_x[-1] + drain[road_i])
        Right_road_design_y.append(Right_road_design_y[-1])
        Right_road_design_x.append(Right_road_design_x[-1])
        Right_road_design_y.append(Right_road_design_y[-1] + drain_cover)
        Right_road_design_x.append(Right_road_design_x[-1] + sidewalk[road_i] + guardrail)
        Right_road_design_y.append(Right_road_design_y[-1] + (sidewalk[road_i] + guardrail) * 0.0137)
        Right_road_design = np.vstack((Right_road_design_y, Right_road_design_x))
    elif type == 'TYPED':  # 要改#右邊沒有任何建築物
        center_y = center_y - kerb
        Right_road_design_x.append(center_x - divisional_island[road_i])
        Right_road_design_y.append(center_y)
        Right_road_design_x.append(Right_road_design_x[-1] - car_lane[road_i])
        Right_road_design_y.append(Right_road_design_y[-1] - car_lane[road_i] * 0.027)
        Right_road_design_x.append(Right_road_design_x[-1] - drain[road_i])
        Right_road_design_y.append(Right_road_design_y[-1])
        Right_road_design_x.append(Right_road_design_x[-1])
        Right_road_design_y.append(Right_road_design_y[-1] + drain_cover)
        Right_road_design_x.append(Right_road_design_x[-1] - sidewalk[road_i])
        Right_road_design_y.append(Right_road_design_y[-1] + sidewalk[road_i] * 0.027)
        Right_road_design = np.vstack((Right_road_design_y, Right_road_design_x))
    elif type == 'TYPEE':
        center_y = center_y - kerb
        Right_road_design_x.append(center_x)
        Right_road_design_y.append(center_y)
        Right_road_design_x.append(Right_road_design_x[-1] + car_lane[road_i])
        Right_road_design_y.append(Right_road_design_y[-1] - car_lane[road_i] * 0.02)
        Right_road_design_x.append(Right_road_design_x[-1] + sidewalk[road_i])
        Right_road_design_y.append(Right_road_design_y[-1])
        Right_road_design_x.append(Right_road_design_x[-1])
        Right_road_design_y.append(Right_road_design_y[-1] + drain_cover)
        Right_road_design = np.vstack((Right_road_design_y, Right_road_design_x))
    n = (np.shape(Right_road_design))[1]
    j = []
    for i in range(1, n):
        if Right_road_design[0][i - 1] == Right_road_design[0][i] and Right_road_design[1][i - 1] == \
                Right_road_design[1][i]:
            j.append(i)
    Right_road_design = np.delete(Right_road_design, [j], 1)
    return Right_road_design
def dis_to_pixel_x(distance):
    pixel = round(distance * (dx + dx2) / 4)
    return pixel
def dis_to_pixel_y(distance):
    pixel = -round(distance * (dy + dy2) / 2)
    return pixel
def abs_dis_to_pixel_x(distance):
    distance = distance - start_level_x
    pixel = distance * (dx + dx2) / 4
    pixel_x = round(pixel + start_grid_x)
    return pixel_x
def abs_dis_to_pixel_y(distance):
    distance = distance - start_level_y
    pixel = -distance * (dy + dy2) / 2
    pixel_y = round(pixel + start_grid_y)
    return pixel_y
def abs_pixel_to_dis_x(pixel):
    pixel = pixel-start_grid_x
    distance = (pixel/(dx+dx2))*4
    distance_x = (distance+start_level_x)
    return distance_x
def abs_pixel_to_dis_y(pixel):
    pixel = pixel-start_grid_y
    distance = (-pixel/(dy+dy2))*2
    distance_y = (distance+start_level_y)
    return distance_y
def station_to_float(station):
    station = station
    j = []
    for i in station:
        if i =='k':
            i = 'k'.upper()
        j.append(i)
    for i in range(0, len(j) - 1):
        if j[i] == 'k'.upper():
            j[i] = 1000
            whereK = i
        elif j[i] == '+':
            wheredel = i
            del j[i]
    station_whereK = whereK - 1
    if station_whereK == 0:
        station_numK = int(j[0])
    elif station_whereK > 0:
        station_numK = j[0]
        for i in j[1:whereK]:
            station_numK += i
    station_numK = int(station_numK)
    station_I = station_numK * j[whereK]
    station_II = j[wheredel + 1:]
    station_III = j[wheredel]
    for i in station_II:
        station_III += i
    station_III = float(station_III)
    station_float = station_I + station_III
    return station_float
def choose_repeat(w):
    w = sorted(w)
    lenw = len(w)
    v = [w[0]]
    for i in range(lenw-1):
        if w[i] != w[i+1]:
            v.append(w[i+1])
    return v
def find_repeat_max(totalnum):
    totalnum = sorted(totalnum)
    len_totalnum = len(totalnum)
    repeat_ever = []
    repeat_value = []
    repeat_times = []
    repeat_max = []
    for i in range(len_totalnum-1):
        if totalnum[i] != totalnum[i+1] and i+1 == len_totalnum-1:
            repeat_ever.append(i)
            repeat_value.append(totalnum[i])
            repeat_ever.append(i+1)
            repeat_value.append(totalnum[i+1])
        elif totalnum[i] == totalnum[i+1] and i+1 == len_totalnum-1:
            repeat_ever.append(i+1)
            repeat_value.append(totalnum[i+1])
        elif totalnum[i] != totalnum[i+1]:
            repeat_ever.append(i)
            repeat_value.append(totalnum[i])
    for i in range(len(repeat_ever)):
        if i ==0:
            repeat_times.append(repeat_ever[i]+1-0)
        else:
            repeat_times.append(repeat_ever[i]-repeat_ever[i-1])
    pos_maxtimes = repeat_times.index(max(repeat_times))
    repeat_max = [repeat_value[pos_maxtimes],repeat_times[pos_maxtimes]]
    return repeat_max
def find_repeat_sec(totalnum):
    totalnum = sorted(totalnum)
    len_totalnum = len(totalnum)
    repeat_ever = []
    repeat_value = []
    repeat_times = []
    repeat_max = []
    for i in range(len_totalnum-1):
        if totalnum[i] != totalnum[i+1] and i+1 == len_totalnum-1:
            repeat_ever.append(i)
            repeat_value.append(totalnum[i])
            repeat_ever.append(i+1)
            repeat_value.append(totalnum[i+1])
        elif totalnum[i] == totalnum[i+1] and i+1 == len_totalnum-1:
            repeat_ever.append(i+1)
            repeat_value.append(totalnum[i+1])
        elif totalnum[i] != totalnum[i+1]:
            repeat_ever.append(i)
            repeat_value.append(totalnum[i])
    for i in range(len(repeat_ever)):
        if i ==0:
            repeat_times.append(repeat_ever[i]+1-0)
        else:
            repeat_times.append(repeat_ever[i]-repeat_ever[i-1])
    pos_maxtimes = repeat_times.index(max(repeat_times))
    max_repeat = repeat_times[pos_maxtimes]
    max_value = repeat_value[pos_maxtimes]
    del repeat_times[pos_maxtimes]
    del repeat_value[pos_maxtimes]
    # pos_maxtimes = repeat_times.index(max(repeat_times))
    # repeat_max = [repeat_value[pos_maxtimes],repeat_times[pos_maxtimes]]

    repeat_times2 = np.array(repeat_times)
    pos_maxtimes = np.where(repeat_times2==max(repeat_times2))
    salvage_value = []
    if np.size(pos_maxtimes) > 1:
        for i in range(int(np.size(pos_maxtimes))):
            j = pos_maxtimes[0][i]
            salvage_value.append(abs(max_value - repeat_value[j]))
        pos_maxj = salvage_value.index(min(salvage_value))
        pos_maxtimes = pos_maxtimes[0][pos_maxj]
        repeat_max = [repeat_value[pos_maxtimes],repeat_times2[pos_maxtimes]]
    else:
        pos_maxtimes = repeat_times.index(max(repeat_times))
        repeat_max = [repeat_value[pos_maxtimes],repeat_times[pos_maxtimes]]
    return repeat_max
def find_dxdy(verti_lst):
    verti_lst=verti_lst
    len_verti_lst = len(verti_lst)
    total_verti=[]
    for i in range(len_verti_lst-1):
        total_verti.append((verti_lst[i+1]-verti_lst[i]))
    repeat_max = find_repeat_max(total_verti)
    repeat_sec = find_repeat_sec(total_verti)
    for i in repeat_sec:
        repeat_max.append(i)
    dxdy = repeat_max
    return dxdy
def grid_position_x(high,width,dx,start_grid_x):
    # 設定總畫布大小
    # 產生畫布img
    shape = (high, width, 3)  # 設定畫布的大小(第一格高、第二格寬) #基本設定
    # 設定網格的間隔
    dx = dx  # 設定寬的間隔  #基本設定
    start_grid_x = start_grid_x#基本設定
    horiz_x = []
    spacing_x = shape[1] / dx
    grid_position_x = []
    #無條件進位 + 轉成正整數
    spacing_x = int(math.ceil(spacing_x * 1)/1.0)
    # for i in range(0,spacing_x):
    #     horiz_x.append(i*dx)
    for i in range(0,spacing_x+1):
        if i == 0:
            horiz_x.append(i*dx)
        elif i%2 ==0:
            horiz_x.append(horiz_x[i-1]+dx)
        elif i%2==1:
            horiz_x.append(horiz_x[i-1]+dx2)
    if horiz_x[-1] > shape[1]:
        horiz_x[-1] = shape[1]
    for i in range(0,spacing_x+1):
        grid_position_x.append(horiz_x[i]+start_grid_x)
    grid_position_x = choose_repeat(grid_position_x)
    # print(f'x={grid_position_x}')
    return grid_position_x
def grid_position_y(high,width,dy,start_grid_y):
    # 設定總畫布大小
    # 產生畫布img
    shape = (high, width, 3)  # 設定畫布的大小(第一格高、第二格寬) #基本設定
    # 設定網格的間隔
    dy = dy  # 設定高的間隔  #基本設定
    start_grid_y = start_grid_y#基本設定
    verti_y = []
    spacing_y = shape[0] / dy
    grid_position_y = []
    #無條件進位 + 轉成正整數
    spacing_y = int(math.ceil(spacing_y * 1)/1.0)

    for i in range(0,spacing_y+1):
        if i == 0:
            verti_y += [i*dy]
           #verti_y.append(i*dy)
        elif i%2 ==0:
            verti_y.append(verti_y[i-1]+dy)
        elif i%2==1:
            verti_y.append(verti_y[i-1]+dy2)
    if verti_y[-1] > shape[0]:
        verti_y[-1] = shape[0]
    for i in range(0,spacing_y+1):
        grid_position_y.append(verti_y[i]+start_grid_y)
    grid_position_y = choose_repeat(grid_position_y)
    # print(f'y={grid_position_y}')
    return grid_position_y
def explan_parameter(v):
#HoughLinesP(image, rho, theta, threshold, minLineLength=None, maxLineGap=None,lines=None) 
#HoughLinesP是OpenCV中用於檢測線條的一個函數。下面是各參數的說明：
#image：要檢測線條的圖像，必須是灰度圖像。
#rho：距離精度，表示像素距離原點的最小單位。
#theta：角度精度，表示弧度值的最小單位。
#threshold：閾值，用於確定檢測到的線條是否有效。只有得票數大於閾值的線條才會被保留。
#minLineLength：最小線段長度，當線段長度小於此值時，將被排除。
#maxLineGap：最大線段間隔，當兩條線段之間的距離小於此值時，將它們視為一個線段。
    #其中，rho和theta的值決定了檢測線條的精度，閾值、最小線段長度和最大線段間隔則決定了檢測到的線條的數量和質量。
#lines是一個輸出參數，用於保存檢測到的線條的坐標。在函數執行後，lines會被填充為一個numpy數組，其中每行都包含檢測到的一條線段的起點和終點坐標。
    #例如，如果檢測到三條線段，則lines會是一個形狀為( 3, 1, 4)的數組，其中第一維表示線段的數量，第二維為1（因為每個線段只有起點和終點），第三維為4
    # 因為每個點有(x, y)兩個坐標）。
    pass
def grid_generate(img,high,width,dy,dy2,dx,dx2,line_width,start_grid_y,start_grid_x):
    # 設定總畫布大小
    # 產生畫布img
    shape = (high, width, 3)  # 設定畫布的大小(第一格高、第二格寬) #基本設定
    origin_img = img
    #origin_img.fill(255) #背景變成白色
    # 設定網格的間隔
    dy = dy  # 設定高的間隔  #基本設定
    dx = dx  # 設定寬的間隔  #基本設定
    line_width = line_width  # 設定網格線寬 #基本設定
    start_grid_y = start_grid_y#基本設定
    start_grid_x = start_grid_x#基本設定
    horiz_x = []
    verti_y = []
    spacing_y = shape[0] / dy
    spacing_x = shape[1] / dx
    #無條件進位 + 轉成正整數
    spacing_y = int(math.ceil(spacing_y * 1)/1.0)
    spacing_x = int(math.ceil(spacing_x * 1)/1.0)
    # for i in range(0,spacing_y):
    #     # verti_x = verti_x+[i*dx]
    #     verti_y += [i*dy]
    for i in range(0,spacing_y+1):
        if i == 0:
            verti_y.append(i*dy)
        elif i%2 ==0:
            verti_y.append(verti_y[i-1]+dy)
        elif i%2==1:
            verti_y.append(verti_y[i-1]+dy2)
    for i in range(0,spacing_x+1):
        if i == 0:
            horiz_x.append(i*dx)
        elif i%2 ==0:
            horiz_x.append(horiz_x[i-1]+dx)
        elif i%2==1:
            horiz_x.append(horiz_x[i-1]+dx2)

    for i in range(0,spacing_y):
        for j in range(0,spacing_x):
            if horiz_x[j]+dx < shape[1] and verti_y[i]+dy < shape[0]:
                grid_img = cv2.rectangle(origin_img ,(horiz_x[j]+start_grid_x,verti_y[i]+start_grid_y) ,(horiz_x[j+1]+start_grid_x,verti_y[i]+dy+start_grid_y),(0,255,0) ,line_width)
            elif verti_y[i]+dy < shape[0]:
                grid_img = cv2.rectangle(origin_img ,(horiz_x[j]+start_grid_x,verti_y[i]+start_grid_y) ,(shape[1]+start_grid_x,verti_y[i]+dy+start_grid_y),(0,255,0) ,line_width)
            else:
                grid_img = cv2.rectangle(origin_img ,(horiz_x[j]+start_grid_x,verti_y[i]+start_grid_y) ,(shape[1]+start_grid_x,shape[0]+start_grid_y),(0,255,0) ,line_width)
    return grid_img
    #return img
def grid_circle(img,high,width,dy,dx,line_width):
    # 設定總畫布大小
    # 產生畫布img
    shape = (high, width, 3)  # 設定畫布的大小(第一格高、第二格寬) #基本設定
    origin_img = img
    # 設定網格的間隔
    dy = dy  # 設定高的間隔  #基本設定
    dx = dx  # 設定寬的間隔  #基本設定
    line_width = line_width  # 設定網格線寬 #基本設定
    horiz_x = []
    verti_y = []
    spacing_y = shape[0] / dy
    spacing_x = shape[1] / dx
    #無條件進位 + 轉成正整數
    spacing_y = int(math.ceil(spacing_y * 1)/1.0)
    spacing_x = int(math.ceil(spacing_x * 1)/1.0)
    for i in range(0,spacing_y):
        # verti_x = verti_x+[i*dx]
        verti_y += [i*dy]
    for i in range(0,spacing_x):
        horiz_x.append(i*dx)

    # 開始繪製網格內點
    if spacing_x > spacing_y:
        for i in range(0, spacing_y):
            point_img = cv2.circle(origin_img, (round((horiz_x[i] + dx) / 2), round((verti_y[i] + dy) / 2)),
                                   (round(((dx / 4) + (dy / 4)) / 2)), (0, 0, 255), line_width)  # 第一版本
            point_img = cv2.circle(origin_img, (round((horiz_x[i] + dx / 2)), round((verti_y[i] + dy / 2))),
                                   (round(((dx / 4) + (dy / 4)) / 2)), (0, 0, 255), line_width)  # 第二版本
    else:
        for i in range(0, spacing_x):
            point_img = cv2.circle(origin_img, (round((horiz_x[i] + dx) / 2), round((verti_y[i] + dy) / 2)),
                                   (round(((dx / 4) + (dy / 4)) / 2)), (0, 0, 255), line_width)  # 第一版本
            point_img = cv2.circle(origin_img, (round((horiz_x[i] + dx / 2)), round((verti_y[i] + dy / 2))),
                                   (round(((dx / 4) + (dy / 4)) / 2)), (0, 0, 255), line_width)  # 第二版本
    return point_img
    #return img
def grid_text(img,high,width,dy,dx,line_width):
    # 設定總畫布大小
    # 產生畫布img
    fontFace = cv2.FONT_HERSHEY_COMPLEX
    shape = (high, width, 3)  # 設定畫布的大小(第一格高、第二格寬) #基本設定
    origin_img = img
    # 設定網格的間隔
    dy = dy  # 設定高的間隔  #基本設定
    dx = dx  # 設定寬的間隔  #基本設定
    line_width = line_width  # 設定網格線寬 #基本設定
    horiz_x = []
    verti_y = []
    spacing_y = shape[0] / dy
    spacing_x = shape[1] / dx
    #無條件進位 + 轉成正整數
    spacing_y = int(math.ceil(spacing_y * 1)/1.0)
    spacing_x = int(math.ceil(spacing_x * 1)/1.0)
    for i in range(0,spacing_y):
        # verti_x = verti_x+[i*dx]
        verti_y += [i*dy]
    for i in range(0,spacing_x):
        horiz_x.append(i*dx)
    j = 0
    for i in range(0,spacing_y):
        k=j+1
        for j in range(0,spacing_x):
            if horiz_x[j]+dx < shape[1]:
                text_img = cv2.putText(origin_img,f'{k*(i)+j+1}' ,(round(horiz_x[j]+dx/2),round(verti_y[i]+dy/2)),fontFace, 1 ,(0,255,0) ,line_width, cv2.LINE_AA)
            elif verti_y[i]+dy < shape[0]:
                text_img = cv2.putText(origin_img,f'{k*(i)+j+1}' ,( round(shape[1]-dx/2) ,round(verti_y[i]+dy/2)),fontFace, 1 ,(0,0,255) ,line_width, cv2.LINE_AA)
            else:
                text_img = cv2.putText(origin_img,f'{k*(i)+j+1}' ,(round(shape[1]-dx/2),round(shape[0]-dy/2)), fontFace, 1 ,(0,0,255) ,line_width, cv2.LINE_AA)
    return text_img
    #return img
def show_xy(event,x,y,flags,userdata):
    #print(event,y,x,flags)
    # 印出相關參數的數值，userdata 可透過 setMouseCallback 第三個參數垂遞給函式
    if event == 0:
        position_img = road_img.copy()                         # 當滑鼠移動時，複製原本的圖片
        cv2.circle(position_img, (x,y), 1, (0,0,0), 1)   # 繪製黑色空心圓
        cv2.imshow('ResetHoughLinesP', position_img)            # 顯示繪製後的影像
    if event == 1:
        yx = f'{y},{x}'
        dots.append([y,x])                          # 記錄座標
        #print(dots)
        cv2.circle(road_img, (x,y), 1, (0, 0, 255), -1)  # 在點擊的位置，繪製圓形
        cv2.putText(road_img, yx, (x, y), cv2.FONT_HERSHEY_PLAIN, 1.0, (0, 0, 255))
        num = len(dots)  # 目前有幾個座標
        if num > 1:  # 如果有兩個點以上
            x1 = dots[num - 2][1]
            y1 = dots[num - 2][0]
            x2 = dots[num - 1][1]
            y2 = dots[num - 1][0]
            cv2.line(road_img, (x1, y1), (x2, y2), (0, 0, 255), 1)  # 取得最後的兩個座標，繪製直線
    return dots
def reselect_y_coordinates(coordinates_lst):
    coordinates_lst = sorted(coordinates_lst)
    bag = []
    result = []
    result.append(coordinates_lst[0])
    for i in coordinates_lst:
        if len(bag) == 0:
            bag.append(i)
        else:
            if i-bag[-1] <=3:               #判斷連號 後面像素-前面像素<=多少算連號 越大越不準
                bag.append(i)
            else:
                if len(bag) >=2:            #若連號的數量>=5
                    result.append(bag[-1])  #把最後一號記錄下來
                bag.clear()
                bag.append(i)
    result.append(i)
    return result
def reselect_x_coordinates(coordinates_lst):
    coordinates_lst = sorted(coordinates_lst)
    bag = []
    result = []
    result.append(coordinates_lst[0])
    for i in coordinates_lst:
        if len(bag) == 0:
            bag.append(i)
        else:
            if i-bag[-1] <=3:
                bag.append(i)
            else:
                if len(bag) >=2:
                    result.append(bag[-1])
                bag.clear()
                bag.append(i)
    result.append(i)
    return result
def grid_ground(img,dots,line_width):
    dots = dots
    num = len(dots)
    line_width = line_width
    for i in range(0,num):
        if i > 0:  # 如果有兩個點以上
            x1 = dots[i - 1][1]
            y1 = dots[i - 1][0]
            x2 = dots[i - 0][1]
            y2 = dots[i - 0][0]
            cv2.line(img, (x1, y1), (x2, y2), (0, 0, 255), line_width)  # 取得最後的兩個座標，繪製直線


'''------下面是從EXCEL讀取道路設定--------'''
wb = load_workbook('excel_road2.xlsx')
print('-------------請認真看說明-------------')
print('請不要開啟excel_road_excel你開著excel檔案就無法複寫，程式就會出錯，請務必注意')
print('-------------請認真看說明-------------\n')
print('使用說明1.輸入路名部分一定要加上M，ex:3-19-30M，請用滑鼠點取原地面線，建議由左至右，只能同一方向點')
print('使用說明2.RuntimeWarning: divide by zero encountered in scalar divide slope = (y2 - y1) / (x2 - x1)')
print('        看到這一行請不要覺得程式出錯，這是斜率為0的警告標語計算上已經處理過，接著會彈出一個圖片視窗就可以開始使用，操作說明如下所述:')
print('使用說明3.圖片出現後請"點擊"滑鼠"左鍵"，建議由左至右點選在「原地面線」上，點數越多精度越高，點完之後會出現紅線，不想再點後就可以關閉視窗(右上角X)')
print('使用說明4.藍色線描繪設計道路完成面以及原地面線(你剛剛點的紅點)，並且綠色網格切在背景圖片的網格越切合計算成果越準')
print('使用說明5.出現上圖後且綠色格線沒有偏離太多背景網格即可關閉視窗，完成第一個斷面挖方面積以及填方面積計算')
print('        一共兩張圖代表一個斷面，反覆操作直至本條道路計算完畢，自動產生excel_road_result.xlsx檔案，開啟即可觀看結果\n')
print('###注意事項1:請確認資料夾的圖片截圖是否為1開始至最後一張，且必須為連號ex:1 2 3 ......')
print('###注意事項2:圖片截圖內容請盡量擷取網格邊就好，圖片上標註之高程以及道路寬度的標註數字都可以不用截到也不會影響計算結果，只需要完整截取網格即可')
print('###注意事項3:本檔案另需搭配使用一個excel檔案作為配件，填寫內容須含\n          樁號 中心高(設計高PGL) 路基 高程最高起始值 中央分隔島 汽車道 側溝 路邊停車 設施人行道 ')
print('###注意事項3.1:ex:(選取道路1-2-50M)樁號2K+620，中心高(設計高PGL):52.799，路基:-0.75，最高起始56，中央分隔島:7.6，汽車道:8，側溝:0.9，路邊停車:0，設施人行道:8.5，type:TYPEA')
print('###注意事項4:以上資訊搭配細設圖使用，TYPE型式也從細設圖可以得到，大部分為A型，BCDE僅少數斷面有，全部採A型計算亦可，差別在於中央分隔島部分會有誤差可再以人力除錯')

now_web = input('請輸入路名')
#輸入目前想要計算的道路名稱
ws = wb[now_web]
colA = ws['A']
numA = np.size(colA)
'''下面五個參數是Output用,上面的是Input'''
excel_station=[]
excel_cut_area=[]
excel_fill_area=[]
excel_need_area=[]
excel_station_float=[]
excel_field_density_layer=[]
excel_survey=[]
excel_subgrade=[]
start_station = int(input('請輸入要第幾張圖開始'))
if start_station == 1:
    end_station = input('請問要直接一路計算到最後一張嗎?\nY/N').upper()
    if end_station == 'N':
        numA = int(input(f'請輸入要到哪一張?\n不可超過{numA-1}'))+1
elif start_station > 1:
    end_station = input('請問要直接一路計算到最後一張嗎?\nY/N').upper()
    if end_station =='N':
        numA = int(input(f'請輸入要到哪一張?\n不可超過{numA-1}'))+1
for road_i in range(start_station,numA):
    station = []
    center_line = []
    subgrade = []
    top_level = []
    divisional_island = []
    car_lane = []
    drain = []
    on_street_parking = []
    sidewalk = []
    type = []
    for i in range(1, numA + 1):
        station.append(ws[f'A{i}'].value)
        center_line.append(ws[f'B{i}'].value)
        subgrade.append(ws[f'C{i}'].value)
        top_level.append(ws[f'D{i}'].value)
        divisional_island.append(ws[f'E{i}'].value)
        car_lane.append(ws[f'F{i}'].value)
        drain.append(ws[f'G{i}'].value)
        on_street_parking.append(ws[f'H{i}'].value)
        sidewalk.append(ws[f'I{i}'].value)
        type.append(ws[f'J{i}'].value)
        if road_i == i - 1:
            print(f'目前樁號={station[road_i]}')
    print(f'station={station}')
    print(f'center_line={center_line}')
    print(f'subgrade={subgrade}')
    print(f'top_level={top_level}')
    print(f'divisional_island={divisional_island}')
    print(f'car_lane={car_lane}')
    print(f'drain={drain}')
    print(f'on_street_parking={on_street_parking}')
    print(f'sidewalk={sidewalk}')
    print(f'type={type}')
    # print(type(station[1]))
    # print(type(center_line[1]))
    # print(type(subgrade[1]))
    # print(type(top_level[1]))
    '''------上面是從EXCEL讀取道路設定--------'''
    ##基本設定Part 1## 設定畫布大小與間隔
    line_width=1
    dots = []   # 記錄座標的空串列
    # 讀取圖像
    image_name = f'./{now_web}/{road_i}.png'
    road_img = cv2.imread(image_name)  #基本設定
    road_shape = np.shape(road_img)
    print(road_shape)
    # 將圖像轉為灰度圖
    gray = cv2.cvtColor(road_img, cv2.COLOR_BGR2GRAY)

    # 執行Canny邊緣檢測
    edges = cv2.Canny(gray, 80, 200, apertureSize=7) #基本設定若切割效果不好需要調整這行

    # 執行霍夫線變換，檢測直線
    minLineLength=10000   #基本設定若切割效果不好需要調整這行
    maxLineGap=125   #基本設定若切割效果不好需要調整這行
    threshold=50       #基本設定若切割效果不好需要調整這行，閥值
    lines = cv2.HoughLinesP(edges, 1, np.pi/180, threshold, minLineLength, maxLineGap, 10)#其中的100是代表取前100條有效直線


    # 在圖像中繪製直線
    for line in lines:
        x1, y1, x2, y2 = line[0]
        cv2.line(road_img, (x1, y1), (x2, y2), (0, 0, 255), 1)

    # 顯示圖像
    # cv2.namedWindow("HoughLinesP",0) #加入這行就可以自由調整視窗為"HoughLinesP"的視窗,後面必須是0,不能是1
    # cv2.imshow('HoughLinesP', road_img)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()

    #設置線段斜率來篩選水平與垂直線
    verti_lst = []
    horiz_lst = []
    verti_lst_coordinate = {}
    horiz_lst_coordinate = {}

    slope = 0
    for line in lines:
        #  將線段的四個點座標分別assign到x1,y1,x2,y2
        x1, y1, x2, y2 = line[0]
        #  計算斜率
        slope = (y2 - y1) / (x2 - x1)
        #  判斷斜率篩選出垂直線與水平線
        if slope == -1 * np.inf: #np.inf代表無限大 -1代表負無限大
            cv2.line(road_img, (x1, y1), (x2, y2), (0, 0, 0), 1)
            #  將垂直線的 x1座標紀錄在verti_lst裡
            verti_lst.append(x1)
            verti_lst_coordinate[f'{x1}'] = line[0]
        if slope == 0:
            cv2.line(road_img, (x1, y1), (x2, y2), (0, 0, 0), 1)
            #  將水平線的 y1座標紀錄在horiz_lst裡
            horiz_lst.append(y1)
            horiz_lst_coordinate[f'{y1}'] = line[0]
    print(sorted(horiz_lst))
    print(sorted(verti_lst))

    # 顯示圖像
    # cv2.imshow('ReHoughLinesP', road_img)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()
    print(f'y座標horiz_lst{reselect_y_coordinates(sorted(horiz_lst))}')
    print(f'x座標verti_lst{reselect_x_coordinates(sorted(verti_lst))}')
    horiz_lst = reselect_y_coordinates(sorted(horiz_lst))
    verti_lst = reselect_x_coordinates(sorted(verti_lst))



    choose_model = input('請輸入計算模式1:用滑鼠點選原地面線\n  2:輸入絕對高程直接產生整條水平直線')
    if choose_model == str(1):
        #將有連號線條保留最後一格後，並重新作畫一次，確認是否乾淨
        road_img = cv2.imread(image_name) #重新讀一次沒有被畫上HoughLinesP的原圖
        for i in horiz_lst:
            if i == horiz_lst_coordinate[f'{i}'][1]:
                x1,y1,x2,y2 = horiz_lst_coordinate[f'{i}']
                cv2.line(road_img, (x1, y1), (x2, y2), (0, 0, 0), 1)
        for i in verti_lst:
            if i == verti_lst_coordinate[f'{i}'][0]:
                x1,y1,x2,y2 = verti_lst_coordinate[f'{i}']
                cv2.line(road_img, (x1, y1), (x2, y2), (0, 0, 0), 1)
        cv2.imshow('ResetHoughLinesP', road_img)
        cv2.setMouseCallback('ResetHoughLinesP', show_xy)  # 設定偵測事件的函式與視窗
        cv2.waitKey(0)
        cv2.destroyAllWindows()
        print(f'dots={dots}')
        excel_field_density_layer.append('用2法才有差值')
        excel_survey.append('同高程才顯示(2法)')
    elif choose_model == str(2):

        elevation = input('輸入絕對高程值')
        elevation = float(elevation)
        #將有連號線條保留最後一格後，並重新作畫一次，確認是否乾淨
        road_img = cv2.imread(image_name) #重新讀一次沒有被畫上HoughLinesP的原圖
        for i in horiz_lst:
            if i == horiz_lst_coordinate[f'{i}'][1]:
                x1,y1,x2,y2 = horiz_lst_coordinate[f'{i}']
                cv2.line(road_img, (x1, y1), (x2, y2), (0, 0, 0), 1)
        for i in verti_lst:
            if i == verti_lst_coordinate[f'{i}'][0]:
                x1,y1,x2,y2 = verti_lst_coordinate[f'{i}']
                cv2.line(road_img, (x1, y1), (x2, y2), (0, 0, 0), 1)
        start_grid_y = horiz_lst[0]
        start_grid_x = verti_lst[0]
        dx_over_dx2_times = 10  # dx跟dx2差多少的話就把dx2變成dx
        dy_over_dy2_times = 8  # dy跟dy2差多少的話就把dx2變成dx
        dx = find_dxdy(verti_lst)[0]
        dx2 = find_dxdy(verti_lst)[2]
        dy = find_dxdy(horiz_lst)[0]
        dy2 = find_dxdy(horiz_lst)[2]
        start_level_y = top_level[road_i]  # from excel get gird_high(start_level)
        start_level_x = 0
        if find_dxdy(verti_lst)[1] - find_dxdy(verti_lst)[3] > dx_over_dx2_times:
            dx2 = dx
        if find_dxdy(horiz_lst)[1] - find_dxdy(horiz_lst)[3] > dy_over_dy2_times:
            dy2 = dy
        if abs(dx - dx2) > 1:
            dx2 = dx
        if abs(dy - dy2) > 1:
            dy2 = dy

        pixel_elevation = abs_dis_to_pixel_y(elevation)
        x1 = verti_lst[0]
        x2 = verti_lst[-1]
        y1 = pixel_elevation
        y2 = y1
        cv2.line(road_img,(x1,y1),(x2,y2),(0,0,255),1)
        cv2.imshow('ResetHoughLinesP', road_img)
        cv2.waitKey(0)
        cv2.destroyAllWindows()
        dots = [[y1,x1], [y2,x2]]
        print(f'dots={dots}')
        if (center_line[road_i]+subgrade[road_i]-elevation) <=0:
            excel_field_density_layer.append(0)
        else:
            excel_field_density_layer.append((center_line[road_i]+subgrade[road_i]-elevation))
        excel_survey.append(elevation)
    else:
        print('不要亂打字!只要輸入1或是2，不要加上.或是任何奇怪的符號\n繼續亂打繼續出錯，你永遠都試不過我已經鎖住了，不然打電話給我')



    ####Part2 基本資料設定####(都已追蹤網格圖片，無須手動更改)
    high=horiz_lst[-1]-horiz_lst[0]
    width=verti_lst[-1]-verti_lst[0]
    start_grid_y = horiz_lst[0]
    start_grid_x = verti_lst[0]
    dx_over_dx2_times = 10     #dx跟dx2差多少的話就把dx2變成dx
    dy_over_dy2_times = 8      #dy跟dy2差多少的話就把dx2變成dx
    # dx=int(math.ceil( ((verti_lst[2]-verti_lst[1]) + (verti_lst[3]-verti_lst[2]))/2 )) #無條件進位
    #dy=int(math.floor(dx/2)) #無條件捨去，讓網格數量最大化
    dx = find_dxdy(verti_lst)[0]
    dx2 = find_dxdy(verti_lst)[2]
    dy = find_dxdy(horiz_lst)[0]
    dy2 = find_dxdy(horiz_lst)[2]
    start_level_y = top_level[road_i] #from excel get gird_high(start_level)
    start_level_x = 0
    kerb=0.35#中央路緣石~~~
    curb=0.15#人行道緣石~~~~
    drain_cover=0.15#水溝蓋~~~
    guardrail = 0.55#大護欄
    if find_dxdy(verti_lst)[1] - find_dxdy(verti_lst)[3] > dx_over_dx2_times:
        print(f'dx={dx},dx2={dx2},次數dx={find_dxdy(verti_lst)[1]},次數dx2={find_dxdy(verti_lst)[3]}')
        dx2 = dx
        print(f'dx2的次數出現少於「dx」10次以上了，所以dx2變成dx')
    if find_dxdy(horiz_lst)[1] - find_dxdy(horiz_lst)[3] > dy_over_dy2_times:
        print(f'dy={dy},dy2={dy2},次數dy={find_dxdy(horiz_lst)[1]},次數dy2={find_dxdy(horiz_lst)[3]}')
        dy2 = dy
        print(f'dy2的次數出現少於「dy」7次以上了，所以dy2變成dy')
    if abs(dx - dx2) >1:
        dx2 = dx
    if abs(dy - dy2) >1:
        dy2 = dy
    print(f'dx={dx},dx2={dx2},次數dx={find_dxdy(verti_lst)[1]},次數dx2={find_dxdy(verti_lst)[3]}')
    print(f'dy={dy},dy2={dy2},次數dy={find_dxdy(horiz_lst)[1]},次數dy2={find_dxdy(horiz_lst)[3]}')
    shape = (high, width, 3)    # 設定畫布的大小(第一格高、第二格寬) #基本設定

    #--------------更改讀取原圖-----------------#
    road_img = cv2.imread(image_name) #重新讀一次沒有被畫上HoughLinesP的原圖
    grid_img = grid_generate(road_img,high,width,dy,dy2,dx,dx2,line_width,start_grid_y,start_grid_x)
    grid_find_position_y = grid_position_y(high,width,dy,start_grid_y)
    grid_find_position_x = grid_position_x(high,width,dx,start_grid_x)
    print(f'grid_find_position_y = {grid_find_position_y}')
    print(f'grid_find_position_x = {grid_find_position_x}')
    ground_line = grid_ground(road_img,dots,line_width)


    #------------畫上道路設計線型--------------------#

    grid_to_elevation_y = excel_y(start_level_y,start_grid_y,grid_find_position_y,dy)
    grid_to_elevation_x = excel_x(start_level_x,start_grid_x,grid_find_position_x,dx)
    num_ele_x = np.size(grid_to_elevation_x)
    #center_pos_x = round(num_ele_x/2)-1
    #center_x = grid_to_elevation_x[center_pos_x]
    center_x = (grid_to_elevation_x[0] + grid_to_elevation_x[num_ele_x-1])/2
    center_y = center_line[road_i]+kerb
    Left_road_draw_point = Left_road_design(center_x,center_y,type[road_i],kerb,curb,drain_cover) #第一項Y第二項X
    num_Left_road_draw_point = (np.shape(Left_road_draw_point))[1]
    Right_road_draw_point = Right_road_design(center_x,center_y,type[road_i],kerb,curb,drain_cover)
    print(f'左邊設計點位={Left_road_draw_point}')
    print(f'左邊設計點位個數={num_Left_road_draw_point}')
    for i in range(0,num_Left_road_draw_point):
        Left_road_draw_point[1][i] = abs_dis_to_pixel_x(Left_road_draw_point[1][i])
        Left_road_draw_point[0][i] = abs_dis_to_pixel_y(Left_road_draw_point[0][i])
        Right_road_draw_point[1][i] = abs_dis_to_pixel_x(Right_road_draw_point[1][i])
        Right_road_draw_point[0][i] = abs_dis_to_pixel_y(Right_road_draw_point[0][i])
    # print(f'左邊設計點位改成像素形式={Left_road_draw_point}')
    # print(f'右邊設計點位改成像素形式={Right_road_draw_point}')

    for i in range(0,num_Left_road_draw_point-1):
        left_x1 = int(Left_road_draw_point[1][i])
        left_y1 = int(Left_road_draw_point[0][i])
        left_x2 = int(Left_road_draw_point[1][i+1])
        left_y2 = int(Left_road_draw_point[0][i+1])
        right_x1 = int(Right_road_draw_point[1][i])
        right_y1 = int(Right_road_draw_point[0][i])
        right_x2 = int(Right_road_draw_point[1][i+1])
        right_y2 = int(Right_road_draw_point[0][i+1])
        cv2.line(road_img, (left_x1, left_y1), (left_x2, left_y2), (255, 0, 0), 1)
        cv2.line(road_img, (right_x1, right_y1), (right_x2, right_y2), (255, 0, 0), 1)
    #繪製圖片，確認用~~可關
    # cv2.imwrite(f'./result/000-000.png', grid_img)
    # cv2.imshow('1234', grid_img)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()

    #---------------讀取重製後的圖-------------------#
    #找尋原地面線的部分!!!!!!!
    #以及找出設計道路的位置~~
    lowerb = np.array([0,0,255])
    upperb = np.array([0,0,255])
    road_lowerb = np.array([255,0,0])
    road_upperb = np.array([255,0,0])
    mask = cv2.inRange(road_img, lowerb, upperb)
    road_mask = cv2.inRange(road_img,road_lowerb,road_upperb)
    ground_position = np.column_stack(np.where(mask == 255))
    road_position = np.column_stack(np.where(road_mask == 255))
    print(f'原地面抓到多少點={np.shape(ground_position)}')
    print(f'道路抓到多少點={np.shape(road_position)}')
    ground_position=np.array(ground_position)
    len_road_position = np.shape(road_position)[0]
    # qweqweqweqwe=np.where(ground_position == 138) #Find all of ground_position
    # print(f'where={qweqweqweqwe}')
    ground_where=[]
    where_road_ground = []
    for i in range(0,len_road_position):
        # if ground_position[:,1] == road_position[i][1]:
        if np.any(ground_position[:,1] == road_position[i][1]):
            # ground_where = np.where(ground_position[:,1] == road_position[i][1])
            ground_where.append(np.where(ground_position[:,1] == road_position[i][1]))
            where_road_ground.append(ground_position[ground_where[i][0][0]])
        else:
            tempground_y=[]
            tempground_x=[]
            if np.any(abs(ground_position[:,1] - road_position[i][1]) <= 10 ):
                slope_where = np.where( abs(ground_position[:,1] - road_position[i][1]) <= 10 )
            for err_i in range(0,np.shape(slope_where)[1]):
                tempground_y.append( abs(ground_position[slope_where[0][err_i]][0] - road_position[i][0]) )
                tempground_x.append( abs(ground_position[slope_where[0][err_i]][1] - road_position[i][1]) )
            err_i=0
            for err_j in range(0,len(tempground_y)-1):
                if tempground_x[err_i]+tempground_y[err_i] > tempground_x[err_j+1]+tempground_y[err_j+1]:
                    err_i = err_j+1
                else:
                    err_i = err_i
            print(ground_position[slope_where[0][err_i]])
            ground_position[slope_where[0][err_i]][1] = road_position[i][1]
            print(ground_position[slope_where[0][err_i]])
            ground_where.append(slope_where[0][err_i])
            print(slope_where[0][err_i])
            where_road_ground.append(ground_position[slope_where[0][err_i]])

    # len_where_road_ground = len_road_position
    # print(f'type,ground_where={np.shape(where_road_ground)[0]}')
    len_where_road_ground = np.shape(where_road_ground)[0]
    print(f'check road and ground is equal road design = {len_where_road_ground}')
    where_road_ground = sorted(where_road_ground,key=lambda x: x[1])
    road_position = sorted(road_position,key=lambda  x: x[1])
    #print(where_road_ground)   #This where_road_ground is not np.array ,so it will post another type with print(where_road_ground)
    where_road_ground=np.array(where_road_ground)
    road_position=np.array(road_position)
    if np.shape(road_position)[0] == np.shape(where_road_ground)[0]:
        print(f'Safe!! road design line is equal ground line,{np.shape(road_position)[0]}')
    else:
        print('check!The for loop of cauculating road desgin line and ground line')
    #此部分在刪除設計道路同個x出現多個y的部分，保留最底的位置而已，如水溝或護欄等......
    repeat_road_design = []
    for i in range(0,len_where_road_ground-1):
        if where_road_ground[i][1] == where_road_ground[i+1][1]:
            repeat_road_design.append(i)
    where_road_ground = np.delete(where_road_ground, [repeat_road_design], 0)
    road_position = np.delete(road_position, [repeat_road_design], 0)
    if np.shape(road_position)[0] == np.shape(where_road_ground)[0]:
        print(f'Safe,where_road_ground=ground line, road_position = road design line,{np.shape(road_position)[0]}')
    else:
        print(f'check!The delete of ground line={np.shape(where_road_ground)[0]} ,and road desgin line={np.shape(road_position)[0]}')
    ########################################################
    cut_area = []
    fill_area = []
    road_position = road_position.astype(float)
    where_road_ground = where_road_ground.astype(float)
    for i in range(0,np.shape(road_position)[0]):
        road_position[i][0] = abs_pixel_to_dis_y(road_position[i][0])
        road_position[i][1] = abs_pixel_to_dis_x(road_position[i][1])
        road_position[i][0] = road_position[i][0] + subgrade[road_i]
        road_position[i][1] = road_position[i][1] + subgrade[road_i]
        where_road_ground[i][0] = abs_pixel_to_dis_y(where_road_ground[i][0])
        where_road_ground[i][1] = abs_pixel_to_dis_x(where_road_ground[i][1])
        if road_position[i][0] - where_road_ground[i][0] > 0:
            fill_area.append(road_position[i][0] - where_road_ground[i][0])
        elif road_position[i][0] - where_road_ground[i][0] < 0:
            cut_area.append(road_position[i][0] - where_road_ground[i][0])
        elif road_position[i][0] - where_road_ground[i][0] ==0 :
            fill_area.append(road_position[i][0] - where_road_ground[i][0])
    #有BUG不知道原因 再確認，np.array儲存類型 原本是int改為float!!!
    sum_cut_area = sum(cut_area)
    sum_fill_area = sum(fill_area)
    sum_need_area = sum_fill_area + sum_cut_area
    sum_cut_area = (sum_cut_area/((dx+dx2)/2))*2
    sum_fill_area = (sum_fill_area/((dx+dx2)/2))*2
    sum_need_area = (sum_need_area/((dx+dx2)/2))*2
    print(f'sum_fill_area={sum_fill_area}')
    print(f'sum_cut_area={sum_cut_area}')
    print(f'sum_need_area={sum_need_area}')
    excel_station.append(station[road_i])
    excel_cut_area.append(sum_cut_area)
    excel_fill_area.append(sum_fill_area)
    excel_need_area.append(sum_need_area)
    excel_station_float.append(station_to_float(station[road_i]))
    excel_subgrade.append(center_line[road_i]+subgrade[road_i])
    #使用Where找尋位置後，再輸入找尋座標值思考過程~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # ground_where = np.where(ground_position[:,1] == 138) #Find Column 0 with every Row!!!!
    # print(f'where={ground_where}')
    # print(f'shape={np.shape(ground_where)}')
    # len_ground_position = len(ground_where[0])
    # print(len_ground_position)
    # for i in range(0,len_ground_position):
    #     print(f'ggggg={ground_position[ground_where[0][i]]}')#ground line
    #思考過程-----------------------------------------------------------------------------------------

    # cv2.imshow('catch',mask)
    # cv2.waitKey(0)     # 按下任意鍵停止
    # cv2.destroyAllWindows()

    len_colorfind = len(ground_position)
    for i in range(len_colorfind):
        if i>0:
            x1 = ground_position[i][1]
            y1 = ground_position[i][0]
            x2 = ground_position[i][1]
            y2 = ground_position[i][0]
            cv2.line(road_img, (x1,y1), (x2,y2), (255,0,0), 1)


    cv2.imshow('test', road_img)
    cv2.waitKey(0)
    cv2.destroyAllWindows()


    #重新產生新的網格
    img = cv2.imread(image_name) #重新讀一次沒有被畫上HoughLinesP的原圖
    new_img = np.zeros(road_shape, np.uint8)
    new_img.fill(255)
    sumcrop_img = {}
    j=0
    for i in range(0,len(horiz_lst)-1):
        k = j+1
        for j in range(0,len(verti_lst)-1):
            crop_img = img[horiz_lst[i]:horiz_lst[i + 1], verti_lst[j]:verti_lst[j + 1]]
            #cv2.imwrite(str(1)+"-"+str(j)+".png",crop_img)
            cv2.imwrite(f'./result/{i+1}-{j+1}.png', crop_img)
            sumcrop_img[f'第{k*(i)+j+1}個'] = crop_img

    j=0
    for i in range(0,len(horiz_lst)-1):
        k = j+1
        for j in range(0,len(verti_lst)-1):
            new_img[horiz_lst[i]:horiz_lst[i + 1], verti_lst[j]:verti_lst[j + 1]] = sumcrop_img[f'第{k*(i)+j+1}個']
    #如果有需要可以看每個被切成微小單位的個別圖片，但對於目前狀況用處不大
    # cv2.imwrite(f'./result/combine_crop.png', new_img)
    # cv2.imshow('combine_crop_img', new_img)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()
    end = time.time()
    print(f'執行時間總共花費{end-start}秒')

'''開始創建xlsx檔案'''
wb2 = Workbook()
ws2 = wb2.active
ws2.title = now_web
ws2['A1'].value = 'Station'
ws2['B1'].value = 'Cut_area'
ws2['C1'].value = 'Fill_area'
ws2['D1'].value = 'Need_area'
ws2['E1'].value = '挖方'
ws2['F1'].value = '填方'
ws2['G1'].value = '需土'
ws2['H1'].value = '路床高程'
ws2['I1'].value = '目前高程'
ws2['K1'].value = '到路床高程需多少公尺'
for i in range(0,len(excel_station)):
    ws2['A'+str(i+2)].value = excel_station[i]
    ws2['B'+str(i+2)].value = excel_cut_area[i]
    ws2['C'+str(i+2)].value = excel_fill_area[i]
    ws2['D'+str(i+2)].value = excel_need_area[i]
    ws2['H'+str(i+2)].value = excel_subgrade[i]
    ws2['I'+str(i+2)].value = excel_survey[i]
    ws2['K'+str(i+2)].value = excel_field_density_layer[i]
for i in range(0,len(excel_station)-1):
    ws2['E'+str(i+2)].value = ((excel_cut_area[i]+excel_cut_area[i+1])/2)*(excel_station_float[i+1]-excel_station_float[i])
    ws2['F'+str(i+2)].value = ((excel_fill_area[i]+excel_fill_area[i+1])/2)*(excel_station_float[i+1]-excel_station_float[i])
    ws2['G'+str(i+2)].value = ((excel_need_area[i]+excel_need_area[i+1])/2)*(excel_station_float[i+1]-excel_station_float[i])
wb2.save(f'./{now_web}/excel_{now_web}road_result.xlsx')
#print(excel_station)
#print(excel_cut_area)
#print(excel_fill_area)
#print(excel_need_area)
#print(excel_need_area[i]*(excel_station_float[i+1]-excel_station_float[i]))